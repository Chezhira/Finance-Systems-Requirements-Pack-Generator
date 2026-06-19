from __future__ import annotations

import csv
from io import BytesIO, StringIO
from itertools import cycle
from pathlib import Path

from finance_requirements_generator.schemas import ControlRiskRow, IntakeAnswers, UATTestCase
from finance_requirements_generator.text_cleanup import clean_fragment, clean_sentence

CONTROL_RISK_HEADERS = [
    "Process Area",
    "Risk Area",
    "Risk Description",
    "Control Objective",
    "Control Activity",
    "Control Type",
    "Frequency",
    "Owner",
    "Evidence Required",
    "System/Data Dependency",
    "Related Requirement ID",
    "Related UAT Case",
    "Residual Risk / Implementation Note",
]

CONTROL_TYPES = [
    "Preventive",
    "Detective",
    "Corrective",
    "Manual",
    "Automated",
    "Semi-automated",
]


def generate_control_risk_matrix(
    process_name: str,
    intake: IntakeAnswers,
    controls: list[str],
    audit_trail_requirements: list[str],
    uat_test_cases: list[UATTestCase],
    functional_requirements: list[str],
    risks_and_dependencies: list[str],
) -> list[ControlRiskRow]:
    rows = []
    control_type_cycle = cycle(CONTROL_TYPES)
    max_rows = max(len(intake.pain_points), len(controls), 1)
    for index in range(max_rows):
        pain_point = _pick(intake.pain_points, index, "Unclear process ownership")
        control = _pick(controls, index, "Control activity requires process owner review")
        audit = _pick(
            audit_trail_requirements,
            index,
            "Evidence trail must be retained for review",
        )
        uat = _pick_case(uat_test_cases, index)
        requirement_id = _identifier(_pick(functional_requirements, index, f"FR-{index + 1:02d}"))
        risk_note = _pick(
            risks_and_dependencies,
            index,
            "Residual risk requires validation during implementation planning",
        )
        rows.append(
            ControlRiskRow(
                process_area=process_name,
                risk_area=clean_fragment(pain_point),
                risk_description=clean_sentence(
                    f"{process_name} may experience {clean_fragment(pain_point).lower()} "
                    "if ownership, data, controls, and evidence are not defined before build."
                ),
                control_objective=clean_sentence(
                    f"Reduce risk from {clean_fragment(pain_point).lower()} through "
                    "clear ownership, evidence, and review criteria."
                ),
                control_activity=clean_sentence(_strip_identifier(control)),
                control_type=next(control_type_cycle),
                frequency=_frequency_for(process_name),
                owner=_owner_for(process_name),
                evidence_required=clean_sentence(_strip_identifier(audit)),
                system_data_dependency=clean_sentence(
                    f"{clean_fragment(intake.erp_platform)} data, required fields, owner status, "
                    "and evidence references must be available for review."
                ),
                related_requirement_id=requirement_id,
                related_uat_case=uat.test_id if uat else "UAT coverage to define",
                residual_risk_implementation_note=clean_sentence(_strip_identifier(risk_note)),
            )
        )
    return rows


def matrix_to_csv_bytes(rows: list[ControlRiskRow]) -> bytes:
    output = StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(CONTROL_RISK_HEADERS)
    for row in rows:
        writer.writerow(_row_values(row))
    return output.getvalue().encode("utf-8-sig")


def export_control_risk_csv(rows: list[ControlRiskRow], output_path: str | Path) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(matrix_to_csv_bytes(rows))
    return path


def matrix_to_xlsx_bytes(rows: list[ControlRiskRow]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "XLSX export requires the openpyxl runtime dependency. "
            'Install the project with `python -m pip install -e ".[dev]"`.'
        ) from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Control Risk Matrix"
    worksheet.append(CONTROL_RISK_HEADERS)
    for row in rows:
        worksheet.append(_row_values(row))

    header_fill = PatternFill("solid", fgColor="16243D")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [22, 24, 46, 44, 44, 18, 18, 24, 42, 46, 20, 18, 48]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_control_risk_xlsx(rows: list[ControlRiskRow], output_path: str | Path) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(matrix_to_xlsx_bytes(rows))
    return path


def _row_values(row: ControlRiskRow) -> list[str]:
    return [
        row.process_area,
        row.risk_area,
        row.risk_description,
        row.control_objective,
        row.control_activity,
        row.control_type,
        row.frequency,
        row.owner,
        row.evidence_required,
        row.system_data_dependency,
        row.related_requirement_id,
        row.related_uat_case,
        row.residual_risk_implementation_note,
    ]


def _pick(items: list[str], index: int, fallback: str) -> str:
    return items[index % len(items)] if items else fallback


def _pick_case(items: list[UATTestCase], index: int) -> UATTestCase | None:
    return items[index % len(items)] if items else None


def _strip_identifier(value: str) -> str:
    if ": " in value:
        return value.split(": ", 1)[1]
    return value


def _identifier(value: str) -> str:
    if ": " in value:
        return value.split(": ", 1)[0]
    return value.split(" ", 1)[0]


def _frequency_for(process_name: str) -> str:
    if "Close" in process_name or "Reconciliation" in process_name:
        return "Each period close"
    if "Payroll" in process_name:
        return "Each payroll cycle"
    return "Each transaction or batch"


def _owner_for(process_name: str) -> str:
    if "Payroll" in process_name:
        return "Payroll Controls Owner"
    if "Close" in process_name:
        return "Financial Controller"
    return f"{process_name} Process Owner"
